import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]

SOURCE_DIR = ROOT / "source"
DATA_DIR = ROOT / "data"

TEAMS_FILE = SOURCE_DIR / "teams.xlsx"
HISTORY_FILE = SOURCE_DIR / "history.xlsx"

FRANCHISES_FILE = DATA_DIR / "franchises.json"
HISTORY_JSON_FILE = DATA_DIR / "history.json"


def clean_text(value):
    """Return trimmed text, or None for an empty cell."""
    if value is None:
        return None

    text = str(value).strip()
    return text if text else None


def to_integer(value, default=0):
    """Convert an Excel value to an integer."""
    if value is None or value == "":
        return default

    return int(float(value))


def to_number(value, default=0.0):
    """Convert an Excel value to a floating-point number."""
    if value is None or value == "":
        return default

    return float(value)


def read_excel_rows(file_path):
    """
    Read the active worksheet and return every populated row
    as a dictionary using the first row as column headers.
    """
    workbook = load_workbook(
        file_path,
        read_only=True,
        data_only=True,
    )

    worksheet = workbook.active
    row_iterator = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(row_iterator)
    except StopIteration:
        workbook.close()
        return []

    headers = [clean_text(value) for value in header_row]
    records = []

    for values in row_iterator:
        record = {}

        for column_index, header in enumerate(headers):
            if not header:
                continue

            value = (
                values[column_index]
                if column_index < len(values)
                else None
            )

            record[header] = value

        if any(value is not None for value in record.values()):
            records.append(record)

    workbook.close()
    return records


def build_franchises(team_rows):
    """
    Build one franchise record per stable TeamId.

    TeamId is permanent.
    TeamName can change between seasons.
    """
    franchises_by_id = {}

    for row in team_rows:
        season = to_integer(row.get("Season"))
        team_id = to_integer(row.get("TeamId"))
        team_name = clean_text(row.get("TeamName"))

        if season == 0 or team_id == 0 or not team_name:
            continue

        franchise_id = str(team_id)

        if franchise_id not in franchises_by_id:
            franchises_by_id[franchise_id] = {
                "franchiseId": franchise_id,
                "currentName": team_name,
                "firstSeason": season,
                "lastSeason": season,
                "historicalNames": [],
                "namesBySeason": [],
            }

        franchise = franchises_by_id[franchise_id]

        franchise["firstSeason"] = min(
            franchise["firstSeason"],
            season,
        )

        franchise["lastSeason"] = max(
            franchise["lastSeason"],
            season,
        )

        franchise["namesBySeason"].append(
            {
                "season": season,
                "teamName": team_name,
            }
        )

    franchises = []

    for franchise in franchises_by_id.values():
        franchise["namesBySeason"].sort(
            key=lambda item: item["season"]
        )

        # The last season in the mapping determines the current name.
        franchise["currentName"] = (
            franchise["namesBySeason"][-1]["teamName"]
        )

        historical_names = []

        for name_record in franchise["namesBySeason"]:
            team_name = name_record["teamName"]

            if team_name not in historical_names:
                historical_names.append(team_name)

        franchise["historicalNames"] = historical_names
        franchises.append(franchise)

    franchises.sort(
        key=lambda item: int(item["franchiseId"])
    )

    return franchises


def build_history(history_rows):
    """Convert historical Excel rows into normalised JSON records."""
    history = []

    for row in history_rows:
        season = to_integer(row.get("Season"))
        team_id = to_integer(row.get("TeamId"))

        if season == 0 or team_id == 0:
            continue

        history.append(
            {
                "season": season,
                "franchiseId": str(team_id),
                "teamName": clean_text(row.get("TeamName")),
                "location": clean_text(row.get("Location")),
                "owner": clean_text(row.get("Owner")),
                "wins": to_integer(row.get("Wins")),
                "losses": to_integer(row.get("Losses")),
                "ties": to_integer(row.get("Ties")),
                "winPct": to_number(row.get("WinPct")),
                "pointsFor": to_number(row.get("PointsFor")),
                "pointsAgainst": to_number(
                    row.get("PointsAgainst")
                ),
                "playoffSeed": to_integer(
                    row.get("PlayoffSeed")
                ),
                "finalRank": to_integer(row.get("FinalRank")),
            }
        )

    history.sort(
        key=lambda item: (
            item["season"],
            item["finalRank"]
            if item["finalRank"] > 0
            else 999,
            int(item["franchiseId"]),
        )
    )

    return history


def validate_source_files():
    """Stop the workflow with a clear message if a file is missing."""
    missing_files = []

    if not TEAMS_FILE.exists():
        missing_files.append(str(TEAMS_FILE))

    if not HISTORY_FILE.exists():
        missing_files.append(str(HISTORY_FILE))

    if missing_files:
        formatted = "\n".join(missing_files)

        raise FileNotFoundError(
            "The following source files are missing:\n"
            f"{formatted}"
        )


def validate_output(franchises, history):
    """Prevent valid data from being replaced by empty JSON files."""
    if not franchises:
        raise ValueError(
            "No franchises were generated from source/teams.xlsx"
        )

    if not history:
        raise ValueError(
            "No historical records were generated "
            "from source/history.xlsx"
        )

    seasons = sorted(
        {record["season"] for record in history}
    )

    print(f"Detected seasons: {seasons}")
    print(f"Franchise count: {len(franchises)}")
    print(f"Historical row count: {len(history)}")


def write_json(file_path, data):
    file_path.write_text(
        json.dumps(
            data,
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def main():
    validate_source_files()
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    team_rows = read_excel_rows(TEAMS_FILE)
    history_rows = read_excel_rows(HISTORY_FILE)

    franchises = build_franchises(team_rows)
    history = build_history(history_rows)

    validate_output(franchises, history)

    write_json(FRANCHISES_FILE, franchises)
    write_json(HISTORY_JSON_FILE, history)

    print(f"Created: {FRANCHISES_FILE}")
    print(f"Created: {HISTORY_JSON_FILE}")


if __name__ == "__main__":
    main()
